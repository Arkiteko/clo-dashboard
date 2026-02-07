"""
Excel report generation for CLO Dashboard using openpyxl.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict, Optional

from src.risk_analytics import (
    compute_warf, compute_diversity_score, compute_hhi,
    compute_portfolio_duration, compute_lien_breakdown,
    compute_single_name_concentration, compute_compliance_status,
)
from src.models import Alert, AlertSeverity


# ── Styling Constants ─────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)


def _style_header_row(ws, row_num, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_df_to_sheet(ws, df, start_row=1):
    """Write a DataFrame to a worksheet with header styling."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    _style_header_row(ws, start_row, len(df.columns))
    # Auto-width columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = max(
            len(str(col_name)) + 4, 14
        )
    return start_row + len(df) + 1


def _write_metric(ws, row, label, value):
    """Write a label-value pair."""
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


# ── Report Generators ─────────────────────────────────────────────────

def generate_global_report(
    df_latest, wh_summary_rows, compliance_rows,
    configs: Optional[Dict] = None,
    alerts: Optional[List[Alert]] = None,
) -> bytes:
    """
    Generate Excel workbook with Global Portfolio data.

    Args:
        df_latest: DataFrame of latest assets across all warehouses
        wh_summary_rows: list of dicts for warehouse comparison table
        compliance_rows: list of dicts for compliance status table
        configs: dict of warehouse_name -> WarehouseConfig (for risk metrics)
        alerts: list of Alert objects
    Returns:
        bytes of the Excel workbook
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Global Portfolio Report").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}")

    row_num = 4
    if not df_latest.empty:
        total_par = df_latest["par_amount"].sum()
        _write_metric(ws, row_num, "Total Funded Exposure", total_par); row_num += 1
        _write_metric(ws, row_num, "Total Assets", len(df_latest)); row_num += 1
        _write_metric(ws, row_num, "Active Warehouses", df_latest["warehouse_source"].nunique()); row_num += 1

        if "market_price" in df_latest.columns and total_par > 0:
            w_price = (df_latest["par_amount"] * df_latest["market_price"]).sum() / total_par
            _write_metric(ws, row_num, "W.Avg Price", round(w_price, 2)); row_num += 1
        if "spread" in df_latest.columns and total_par > 0:
            w_spread = (df_latest["par_amount"] * df_latest["spread"]).sum() / total_par
            _write_metric(ws, row_num, "W.Avg Spread (bps)", round(w_spread, 0)); row_num += 1

        # Risk metrics
        row_num += 1
        ws.cell(row=row_num, column=1, value="Risk Metrics").font = SUBTITLE_FONT; row_num += 1
        warf = compute_warf(df_latest)
        diversity = compute_diversity_score(df_latest)
        duration_data = compute_portfolio_duration(df_latest)
        issuer_hhi = compute_hhi(df_latest, "issuer_name")
        industry_hhi = compute_hhi(df_latest, "industry_gics") if "industry_gics" in df_latest.columns else 0.0

        _write_metric(ws, row_num, "WARF", round(warf, 0)); row_num += 1
        _write_metric(ws, row_num, "Diversity Score", round(diversity, 1)); row_num += 1
        _write_metric(ws, row_num, "W.Avg Duration", round(duration_data["weighted_avg_duration"], 2)); row_num += 1
        _write_metric(ws, row_num, "Portfolio DV01", round(duration_data["portfolio_dv01"], 0)); row_num += 1
        _write_metric(ws, row_num, "Issuer HHI", round(issuer_hhi, 4)); row_num += 1
        _write_metric(ws, row_num, "Industry HHI", round(industry_hhi, 4)); row_num += 1

    # Sheet 2: Warehouse Comparison
    if wh_summary_rows:
        ws2 = wb.create_sheet("Warehouse Comparison")
        df_wh = pd.DataFrame(wh_summary_rows)
        _write_df_to_sheet(ws2, df_wh)

    # Sheet 3: Compliance
    if compliance_rows:
        ws3 = wb.create_sheet("Compliance")
        df_comp = pd.DataFrame(compliance_rows)
        _write_df_to_sheet(ws3, df_comp)

    # Sheet 4: Risk Metrics by Warehouse
    if not df_latest.empty and configs:
        ws_risk = wb.create_sheet("Risk Metrics")
        risk_rows = []
        for wh_name, group in df_latest.groupby("warehouse_source"):
            cfg = configs.get(wh_name)
            dur = compute_portfolio_duration(group)
            lien = compute_lien_breakdown(group)
            sn = compute_single_name_concentration(group)
            risk_rows.append({
                "Warehouse": wh_name,
                "WARF": round(compute_warf(group), 0),
                "Diversity Score": round(compute_diversity_score(group), 1),
                "W.Avg Duration": round(dur["weighted_avg_duration"], 2),
                "Portfolio DV01": round(dur["portfolio_dv01"], 0),
                "Issuer HHI": round(compute_hhi(group, "issuer_name"), 4),
                "Industry HHI": round(compute_hhi(group, "industry_gics"), 4) if "industry_gics" in group.columns else 0,
                "1L %": f"{lien['1L_pct']:.1%}",
                "2L %": f"{lien['2L_pct']:.1%}",
                "Unsecured %": f"{lien['unsecured_pct']:.1%}",
                "Max Single Name %": f"{sn['max_single_issuer_pct']:.1%}",
                "Max Single Name": sn["max_single_issuer_name"],
            })
        if risk_rows:
            _write_df_to_sheet(ws_risk, pd.DataFrame(risk_rows))

    # Sheet 5: Asset Detail
    if not df_latest.empty:
        ws4 = wb.create_sheet("Assets")
        asset_cols = [c for c in ["asset_id", "issuer_name", "warehouse_source", "par_amount",
                                   "market_price", "spread", "rating_moodys", "rating_sp",
                                   "lien_type", "maturity_date", "industry_gics", "country",
                                   "is_defaulted"] if c in df_latest.columns]
        _write_df_to_sheet(ws4, df_latest[asset_cols])

    # Sheet 6: Alerts Summary
    if alerts:
        ws_alerts = wb.create_sheet("Alerts")
        alert_rows = []
        for a in alerts:
            alert_rows.append({
                "Severity": a.severity.value,
                "Warehouse": a.warehouse,
                "Category": a.category,
                "Alert": a.title,
                "Detail": a.detail,
                "Current Value": a.current_value,
                "Threshold": a.threshold_value,
            })
        if alert_rows:
            _write_df_to_sheet(ws_alerts, pd.DataFrame(alert_rows))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_warehouse_report(
    df_wh, config, warehouse_name,
    alerts: Optional[List[Alert]] = None,
) -> bytes:
    """
    Generate Excel workbook for a single warehouse.
    """
    wb = Workbook()

    # Sheet 1: Metrics
    ws = wb.active
    ws.title = "Metrics"
    ws.cell(row=1, column=1, value=f"Warehouse Report: {warehouse_name}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}")
    ws.cell(row=3, column=1, value=f"Type: {config.warehouse_type}")

    row_num = 5
    if not df_wh.empty:
        funded = df_wh["par_amount"].sum()
        _write_metric(ws, row_num, "Funded Exposure", funded); row_num += 1
        _write_metric(ws, row_num, "Asset Count", len(df_wh)); row_num += 1
        _write_metric(ws, row_num, "Max Facility", config.max_facility_amount); row_num += 1
        _write_metric(ws, row_num, "Advance Rate", config.advance_rate); row_num += 1

        if funded > 0 and "market_price" in df_wh.columns:
            w_price = (df_wh["par_amount"] * df_wh["market_price"]).sum() / funded
            _write_metric(ws, row_num, "W.Avg Price", round(w_price, 2)); row_num += 1
        if funded > 0 and "spread" in df_wh.columns:
            w_spread = (df_wh["par_amount"] * df_wh["spread"]).sum() / funded
            _write_metric(ws, row_num, "W.Avg Spread (bps)", round(w_spread, 0)); row_num += 1

        # Risk metrics
        row_num += 1
        ws.cell(row=row_num, column=1, value="Risk Metrics").font = SUBTITLE_FONT; row_num += 1
        warf = compute_warf(df_wh)
        diversity = compute_diversity_score(df_wh)
        dur = compute_portfolio_duration(df_wh)
        lien = compute_lien_breakdown(df_wh)
        sn = compute_single_name_concentration(df_wh)

        _write_metric(ws, row_num, "WARF", round(warf, 0)); row_num += 1
        _write_metric(ws, row_num, "Diversity Score", round(diversity, 1)); row_num += 1
        _write_metric(ws, row_num, "W.Avg Duration", round(dur["weighted_avg_duration"], 2)); row_num += 1
        _write_metric(ws, row_num, "Portfolio DV01", round(dur["portfolio_dv01"], 0)); row_num += 1
        _write_metric(ws, row_num, "Issuer HHI", round(compute_hhi(df_wh, "issuer_name"), 4)); row_num += 1
        _write_metric(ws, row_num, "1L %", f"{lien['1L_pct']:.1%}"); row_num += 1
        _write_metric(ws, row_num, "2L %", f"{lien['2L_pct']:.1%}"); row_num += 1
        _write_metric(ws, row_num, "Unsecured %", f"{lien['unsecured_pct']:.1%}"); row_num += 1
        _write_metric(ws, row_num, "Max Single Name", f"{sn['max_single_issuer_name']} ({sn['max_single_issuer_pct']:.1%})"); row_num += 1

    # Sheet 2: Assets
    if not df_wh.empty:
        ws2 = wb.create_sheet("Assets")
        asset_cols = [c for c in ["asset_id", "issuer_name", "par_amount", "market_price",
                                   "spread", "rating_moodys", "rating_sp", "lien_type",
                                   "maturity_date", "industry_gics", "country",
                                   "is_defaulted"] if c in df_wh.columns]
        _write_df_to_sheet(ws2, df_wh[asset_cols])

    # Sheet 3: Lien Breakdown
    if not df_wh.empty:
        lien_data = compute_lien_breakdown(df_wh)
        if not lien_data["breakdown_table"].empty:
            ws_lien = wb.create_sheet("Lien Breakdown")
            _write_df_to_sheet(ws_lien, lien_data["breakdown_table"])

    # Sheet 4: Alerts
    if alerts:
        ws_alerts = wb.create_sheet("Alerts")
        alert_rows = []
        for a in alerts:
            alert_rows.append({
                "Severity": a.severity.value,
                "Category": a.category,
                "Alert": a.title,
                "Detail": a.detail,
                "Current Value": a.current_value,
                "Threshold": a.threshold_value,
            })
        if alert_rows:
            _write_df_to_sheet(ws_alerts, pd.DataFrame(alert_rows))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_stress_report(stress_results, scenario_rows, warehouse_name) -> bytes:
    """
    Generate Excel workbook with stress test results.

    Args:
        stress_results: StressResults object from run_all_scenarios
        scenario_rows: list of dicts with scenario breakdown
        warehouse_name: warehouse identifier
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Stress Test Report: {warehouse_name}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}")

    _write_metric(ws, 4, "Total Par", stress_results.total_par)
    _write_metric(ws, 5, "Base OC Ratio", f"{stress_results.base_oc:.2%}")
    _write_metric(ws, 6, "Stressed OC Ratio", f"{stress_results.stressed_oc:.2%}")
    _write_metric(ws, 7, "OC Trigger", f"{stress_results.oc_trigger:.0%}")
    _write_metric(ws, 8, "OC Breach", "YES" if stress_results.oc_breach else "No")
    _write_metric(ws, 9, "Total Stressed Loss", stress_results.total_stressed_loss)
    _write_metric(ws, 10, "Stressed CCC %", f"{stress_results.stressed_ccc_pct:.1%}")
    _write_metric(ws, 11, "CCC Breach", "YES" if stress_results.ccc_breach else "No")

    # Sheet 2: Scenario Breakdown
    if scenario_rows:
        ws2 = wb.create_sheet("Scenarios")
        df_scen = pd.DataFrame(scenario_rows)
        _write_df_to_sheet(ws2, df_scen)

    # Sheet 3+: Per-scenario asset detail
    for s in stress_results.scenarios:
        if s.asset_level is not None and not s.asset_level.empty:
            safe_name = s.name[:28]  # Excel sheet name max 31 chars
            ws_s = wb.create_sheet(safe_name)
            _write_df_to_sheet(ws_s, s.asset_level)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
